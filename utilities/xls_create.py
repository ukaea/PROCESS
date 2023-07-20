#!/usr/bin/env python

"""
Append summary data to a spreadsheet data_summary.xlsx or name specified.
Michael Kovari

The descriptions are output only once, at the head of each worksheet.
The variable names are output each time the utility is used, so the user can check that they haven't changed.
Some clever person could include an automated check instead.  If this is done, it is still a good idea to leave a blank line -
When Excel plots a graph, it handles blank lines by simply leaving a gap in the line joining the points.

Data source: MFILE.DAT or as specified
Configuration file: xls.conf (Not required.  Default config file will be written if it is not present.)
Output file data_summary.xlsx or as specified

I have used openPyXL.
Charts in existing workbooks will be lost.
If you want to add a chart to the workbook, you will need to save a copy.

You can now specify the sheet to which the data will be appended, with argument '-n'.
The sheet will be created if it does not exist.
"""

import os
import argparse
import process.io.mfile as mf
from openpyxl import Workbook, load_workbook


def append_line(spreadsheet, custom_keys, mfile_data):
    """Function to add data to a spreadsheet using MFILE"""

    num_scans = int(mfile_data.data["isweep"].get_scan(-1))

    # If isweep does not appear in MFILE, then the mfile_data.data function will
    # catch the error, and num_scans will be 0
    if num_scans == 0:
        num_scans = 1
    print("num_scans", num_scans)

    val_keys = []
    mfile_keys = mfile_data.data.keys()

    try:
        wb = load_workbook(spreadsheet)
    except Exception:
        wb = Workbook()

    # Use the worksheet with the specified name
    if args.n:
        try:
            ws = wb.get_sheet_by_name(args.n)
        except Exception:
            ws = wb.create_sheet(title=args.n)
            print("New sheet created:", args.n)
    else:
        # Use whichever worksheet is active
        ws = wb.active

    var_descriptions = [""]
    var_names = [""]
    for key in custom_keys:
        if key in mfile_keys:
            var_description = mfile_data.data[key].var_description.replace(" ", "_")
            var_descriptions = var_descriptions + [var_description]
            val_keys.append(key)
            var_names = var_names + [key]
        else:
            # The variable does not appear in the MFILE.  Leave a space
            var_description = ""
            var_descriptions = var_descriptions + [var_description]
            val_keys.append(key)
            var_names = var_names + [key]

    # Print the descriptions only once on each sheet
    try:
        if ws.max_row == 1:
            ws.append(var_descriptions)
    except Exception:
        # Don't write the header line
        pass

    ws.append(var_names)

    # Write rows of values. One row for each scan.
    for num in range(num_scans):
        new_row = [""]
        value = ""
        for vkey in val_keys:
            if vkey in header_variables:
                # These are found only once at the top of the MFILE.
                value = mfile_data.data[vkey].get_scan(-1)
            else:
                # These are present in the MFILE for each scan point
                value = mfile_data.data[vkey].get_scan(num + 1)
            new_row = new_row + [value]
        ws.append(new_row)

    # Save the spreadsheet
    wb.save(spreadsheet)
    print("Data appended to worksheet", ws.title)


# -----------------------------------------------------------------------------


if __name__ == "__main__":

    # Setup command line arguments
    parser = argparse.ArgumentParser(
        description="Append summary data to a spreadsheet data_summary.xlsx or name specified."
        + "  Config file xls.conf optional"
    )

    parser.add_argument(
        "-p", metavar="p", type=str, nargs="+", help="add new variables to the output"
    )

    parser.add_argument("-f", metavar="f", type=str, help="File to read as MFILE.DAT")

    parser.add_argument(
        "-x", metavar="x", type=str, help="Workbook (.xlsx) file to append to"
    )

    parser.add_argument(
        "-n",
        metavar="n",
        type=str,
        help="Use the worksheet (tab) with specified name.  Sheet will be created if it does not exist.",
    )

    parser.add_argument(
        "--defaults", help="run with default params", action="store_true"
    )

    parser.add_argument("--reset-config", help="Reset xls.conf", action="store_true")

    args = parser.parse_args()

    default_variables = [
        "runtitle",
        "username",
        "date",
        "iscan",
        "rmajor",
        "aspect",
        "powfmw",
        "bt",
        "beta",
        "te",
        "te0",
        "dene",
        "ralpne",
        "dnz",
        "pradmw",
        "pdivt",
        "hfact",
        "pinjmw",
        "tburn",
        "ttarget",
        "fmom",
        "qtargetcomplete",
        "qtarget",
        "totalpowerlost",
        "vburn",
        "vsstt",
        "bore",
        "ohcth",
        "tfcth",
        "tmarg",
        "sig_hoop",
        "sig_axial",
        "sig_axial",
        "tesep",
        "nesep",
        "ieped",
        "teped",
        "neped",
    ]
    # Also append these variables by default:
    header_variables = [
        "procver",
        "date",
        "time",
        "username",
        "runtitle",
        "tagno",
        "isweep",
        "nsweep",
    ]

    # If user has specified an MFILE file that isn't MFILE.DAT pass the filename to
    # MFILE() class.
    if args.f:
        M = mf.MFile(filename=args.f)
    else:
        M = mf.MFile()

    # If user has specified a workbook file that isn't data_summary.xlsx
    if args.x:
        spreadsheet = args.x
    else:
        spreadsheet = "data_summary.xlsx"

    print("Workbook name", spreadsheet)

    # Get files in current directory to check for the config file.
    current_directory = os.listdir(".")
    if "xls.conf" not in current_directory or args.reset_config:
        print("Configuration file xls.conf not found in the local directory")
        conf_file = open("xls.conf", "a")
        for item in default_variables:
            conf_file.write(item + "\n")
        conf_file.close()
        print("A default configuration file xls.conf has been written ")

    # Read the config file.
    INPUT_CONFIG = mf.read_mplot_conf("xls.conf")

    # If the user added new parameters in the command line add them to the
    # INPUT_CONFIG list to pass to make_plot_dat()
    if args.p:
        conf_file = open("xls.conf", "a")
        for item in args.p:
            if item not in INPUT_CONFIG:
                conf_file.write(item + "\n")
        conf_file.close()

    INPUT_CONFIG = mf.read_mplot_conf("xls.conf")
    print(INPUT_CONFIG)

    append_line(spreadsheet, INPUT_CONFIG, M)
